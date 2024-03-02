from pathlib import Path
from typing import Optional
from openpyxl import load_workbook
import csv
import json
from datetime import datetime


def xlsx_to_pdf(xlsx_file_path: Path, csv_file_path: Optional[Path] = None):
    csv_file_path = csv_file_path or str(xlsx_file_path) + '.csv'
    # file = pd.read_excel(xlsx_file_path)
    # file.to_csv(csv_file_path)
    wb = load_workbook(xlsx_file_path)
    sh = wb.active
    with open(csv_file_path, 'w', newline="") as file_handle:
        csv_writer = csv.writer(file_handle)
        for row in sh.iter_rows():  # generator; was sh.rows
            csv_writer.writerow([cell.value for cell in row])
    return csv_file_path


def alfabank_operations_to_csv(json_file_path: Path, csv_file_path: Optional[Path] = None):
    with open(json_file_path) as f:
        operations = json.load(f)

    csv_file_path = csv_file_path or str(json_file_path) + '.csv'

    with open(csv_file_path, 'w', newline="") as csv_file:
        fieldnames = ["id", "dateTime", "logoUrl", "title", "amount", "comment", "mcc", "category_id", "category_name", "direction", "loyalty_amount", "status"]
        csv_writer = csv.DictWriter(csv_file, fieldnames=fieldnames)

        csv_writer.writeheader()
        for operation in operations:
            csv_writer.writerow(dict(zip(fieldnames, [
                operation['id'],
                datetime.fromisoformat(operation['dateTime']).isoformat(),
                operation['logoUrl'],
                operation['title'],
                operation['amount']['value'] / operation['amount']['minorUnits'],
                operation['comment'],
                operation['mcc'],
                operation['category']['id'],
                operation['category']['name'],
                operation['direction'].lower() if operation['direction'] else None,
                operation['loyalty']['amount']['value'] / operation['loyalty']['amount']['minorUnits'] if operation['loyalty'] else None,
                operation['status'].lower() if operation['status'] else None
            ])))

    return csv_file_path


# def import_to_database():
#
